import openpyxl
import smtplib
import sys

my_mail=" " #enter your email here
my_pass=" " #enter the pass

workbook=openpyxl.load_workbook("C:\\Users\\aulee\\OneDrive\\Documents\\MyFile.xlsx") #Loading my email container file
sheet=workbook["Sheet1"]

flag=0

for i in range(1,10):
    if str(sheet.cell(1,i).value)=="Email":     #Identifying column containing the Email IDs
        flag=1
        print("Found")
        break;
if flag==0:#Exits if no email found
        sys.exit(0)


reciever=[]
x=2
val=sheet.cell(x,i).value
while(val!=None):
    val=sheet.cell(row=x, column=i).value
    reciever.append(val)                        #Storing the email IDs found in a list
    x+=1



message='''SUBJECT: Test Mail

Hey! This is an self generated mail. If you have recieved this mail then my program worked fine :')

Do not forget to send me a Hi,tho.
Long time. :')

Regards

Auleen
'''

connection=smtplib.SMTP('smtp.gmail.com',587)   #Setting up SMTP
connection.ehlo()                               #Making Connection
connection.starttls()                           #Ensuring Encryption
connection.login(my_mail,my_pass)               #Login
try:
    for email_id in reciever:                         #Accessing the email Ids
        connection.sendmail(my_mail,email_id,message) #Sending Mails
except:
    print("Worked Fine")
connection.quit()

#---------------------------------------------
#Add Feature to input mail text from some Files.

#----------------------------------------------
